import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Change file name
month_year = "Oct_2024"
input_file = 'Register_Timekeeping_Report_2024-09-26_00_00GMT_to_2024-10-29_23_59GMT.xlsx'

def format_excel_file(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        column = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    # Define background colors
    light_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    dark_grey = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Apply alternating background colors to Clock In/Clock Out pairs
    is_light_grey = True
    current_employee = None
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        employee = row[2].value
        event = row[3].value
        
        if employee != current_employee:  # Reset color pattern for new employee
            current_employee = employee
            is_light_grey = True
        
        if event in ["Clock In", "Clock Out"]:
            fill_color = light_grey if is_light_grey else dark_grey
            for cell in row:
                cell.fill = fill_color
            if event == "Clock Out":  # Toggle color after "Clock Out"
                is_light_grey = not is_light_grey
        
        if event == "Total Hours":
            row[4].fill = yellow
    
    wb.save(file_path)

def calculate_work_hours(input_file):
    df = pd.read_excel(input_file, header=None)
    
    # Select necessary columns
    columns_to_select = [0, 1, 5, 6]
    df = df.iloc[:, columns_to_select]
    df.columns = ['Date', 'Time', 'Employee', 'Event']
    
    # Strip unwanted substring from time
    df['Time'] = df['Time'].str[:8]
    
    # Combine 'Date' and 'Time' columns to create a datetime column
    df['Datetime'] = pd.to_datetime(
        df['Date'].astype(str) + ' ' + df['Time'], 
        format='%Y-%m-%d %H:%M:%S', 
        errors='coerce'
    )
    
    # Filter out invalid rows
    df = df[df['Datetime'].notna()]
    
    # Group by Employee
    employee_groups = df.groupby('Employee')
    
    # Create a folder for the month and year
    folder_name = month_year
    os.makedirs(folder_name, exist_ok=True)
    
    # Prepare a combined DataFrame for all employees
    combined_df = pd.DataFrame()
    
    for employee, group in employee_groups:
        # Sort group by Datetime
        group = group.sort_values(by='Datetime').reset_index(drop=True)
        
        # Initialize the 'Work Hours' column
        group['Work Hours'] = None
        
        # Calculate work hours within each employee group
        work_hours = []
        for i in range(1, len(group)):
            if group.iloc[i-1]['Event'] == 'Clock In' and group.iloc[i]['Event'] == 'Clock Out':
                # Calculate time difference
                work_time = group.iloc[i]['Datetime'] - group.iloc[i-1]['Datetime']
                # Convert timedelta to hr:min format
                hours, remainder = divmod(work_time.total_seconds(), 3600)
                minutes = remainder // 60
                work_hours.append(f"{int(hours):02}:{int(minutes):02}")
            else:
                work_hours.append(None)
        
        # Add calculated work hours to the group
        group['Work Hours'] = [None] + work_hours  # Ensure alignment with rows
        
        # Calculate total hours for the employee as timedelta objects
        total_seconds = group['Work Hours'].dropna().apply(
            lambda x: int(x.split(':')[0]) * 3600 + int(x.split(':')[1]) * 60
        ).sum()
        
        # Convert total_seconds back to hr:min format
        total_hours = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}"
        
        # Create a total hours row with the same number of columns as the original DataFrame
        total_hours_row = group.iloc[0:1].copy()  # Copy the first row's structure
        total_hours_row.iloc[0] = np.nan  # Use np.nan instead of '' to avoid dtype conflict
        total_hours_row['Event'] = 'Total Hours'
        total_hours_row['Work Hours'] = total_hours
        
        # Concatenate the group with its total hours row
        employee_df = pd.concat([group, total_hours_row])

        # Save individual employee file
        employee_df = employee_df.drop(columns=['Datetime'], errors='ignore') 
        employee_file = os.path.join(folder_name, f"{employee.replace(' ', '_')}_{month_year}.xlsx")
        employee_df.to_excel(employee_file, index=False)
        format_excel_file(employee_file)
        
        # Add employee data to the combined DataFrame
        combined_df = pd.concat([combined_df, employee_df, pd.DataFrame([[''] * len(group.columns)], columns=group.columns)])
    
    # save combined file
    combined_df = combined_df.drop(columns=['Datetime'], errors='ignore')
    combined_file = os.path.join(folder_name, f"All_{month_year}.xlsx")
    combined_df.to_excel(combined_file, index=False)
    format_excel_file(combined_file)
    
    print(f"All files have been saved in the folder: {folder_name}")

calculate_work_hours(input_file)